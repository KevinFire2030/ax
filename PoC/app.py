import json
import os
import re
import time
import zipfile
from datetime import datetime
from pathlib import Path

import openpyxl
import requests
import streamlit as st
from docx import Document
from dotenv import load_dotenv

load_dotenv()

BASE_DIR = Path(__file__).resolve().parent
WORK_DIR = BASE_DIR / "runs"
WORK_DIR.mkdir(exist_ok=True)

RUBRIC_COLUMNS = {
    "과제 목적 명확성": 11,
    "전략 목표 연계": 12,
    "정량적 효과 제시": 13,
    "기대효과 구체성": 14,
    "추진일정 상세도": 15,
    "투입 자원 명확성": 16,
    "리스크 인식": 17,
    "Pain Point 분석": 18,
    "기술적 차별화": 19,
    "프로세스 혁신": 20,
    "확산 가능성/실적": 21,
    "활용 범위": 22,
    "데이터 종류/출처": 23,
    "데이터 품질/처리": 24,
    "AI 기술 명시": 25,
    "기술 선정 이유": 26,
    "기술 활용 수준": 27,
    "핵심 기능 구체성": 28,
    "아키텍처/파이프라인": 29,
}

RUBRIC = {
    "과제 목적 명확성": {"우수": 5, "양호": 3, "보통": 1, "미흡": 0},
    "전략 목표 연계": {"우수": 5, "양호": 3, "보통": 1, "미흡": 0},
    "정량적 효과 제시": {"우수": 8, "양호": 5, "보통": 2, "미흡": 0},
    "기대효과 구체성": {"우수": 7, "양호": 4, "보통": 2, "미흡": 0},
    "추진일정 상세도": {"우수": 5, "양호": 3, "보통": 1, "미흡": 0},
    "투입 자원 명확성": {"우수": 5, "양호": 3, "보통": 1, "미흡": 0},
    "리스크 인식": {"우수": 5, "양호": 3, "보통": 1, "미흡": 0},
    "Pain Point 분석": {"우수": 10, "양호": 6, "보통": 3, "미흡": 0},
    "기술적 차별화": {"우수": 5, "양호": 3, "보통": 1, "미흡": 0},
    "프로세스 혁신": {"우수": 5, "양호": 3, "보통": 1, "미흡": 0},
    "확산 가능성/실적": {"우수": 5, "양호": 3, "보통": 1, "미흡": 0},
    "활용 범위": {"우수": 5, "양호": 3, "보통": 1, "미흡": 0},
    "데이터 종류/출처": {"우수": 5, "양호": 3, "보통": 1, "미흡": 0},
    "데이터 품질/처리": {"우수": 5, "양호": 3, "보통": 1, "미흡": 0},
    "AI 기술 명시": {"우수": 3, "양호": 2, "보통": 1, "미흡": 0},
    "기술 선정 이유": {"우수": 2, "양호": 1, "미흡": 0},  # 보통 없음
    "기술 활용 수준": {"우수": 5, "양호": 3, "보통": 1, "미흡": 0},
    "핵심 기능 구체성": {"우수": 5, "양호": 3, "보통": 1, "미흡": 0},
    "아키텍처/파이프라인": {"우수": 5, "양호": 3, "보통": 1, "미흡": 0},
}

METRIC_TO_RUBRIC = {
    "목표적합성": ["과제 목적 명확성", "전략 목표 연계"],
    "경제효과성": ["정량적 효과 제시", "Pain Point 분석"],
    "실행가능성": ["추진일정 상세도", "투입 자원 명확성", "리스크 인식"],
    "구체성": ["기대효과 구체성", "핵심 기능 구체성"],
    "효율성": ["정량적 효과 제시", "프로세스 혁신"],
    "혁신성": ["기술적 차별화", "프로세스 혁신"],
    "확상성": ["확산 가능성/실적", "활용 범위"],
    "데이터수집": ["데이터 종류/출처"],
    "데이터처리": ["데이터 품질/처리", "아키텍처/파이프라인"],
    "AI기술": ["AI 기술 명시", "기술 선정 이유", "기술 활용 수준"],
}

SYSTEM_PROMPT = """당신은 MX 과제제안서 블라인드 심사관입니다.
아래 19개 세부항목 각각을 반드시 우수/양호/보통/미흡 중 하나로 판정하세요.
(단, '기술 선정 이유' 항목은 보통 없음: 우수/양호/미흡만 허용)

세부항목:
과제 목적 명확성, 전략 목표 연계, 정량적 효과 제시, 기대효과 구체성,
추진일정 상세도, 투입 자원 명확성, 리스크 인식, Pain Point 분석,
기술적 차별화, 프로세스 혁신, 확산 가능성/실적, 활용 범위,
데이터 종류/출처, 데이터 품질/처리, AI 기술 명시, 기술 선정 이유,
기술 활용 수준, 핵심 기능 구체성, 아키텍처/파이프라인

출력은 반드시 JSON만 반환:
{
  \"rubric_levels\": {각 세부항목: \"우수|양호|보통|미흡\"},
  \"opinion\": \"종합의견(장점/단점, 불합격 사유가 있으면 납득 가능하게)\"
}
"""

REJECTION_PROMPT = """당신은 엄정한 심사관입니다.
아래 정보를 바탕으로 '불합격 사유'를 한국어로 상세히 작성하세요.
요구사항:
- 납득 가능한 근거 중심
- 점수 하락 원인(경영성과/혁신성·확산성/AI·데이터) 분리
- 무엇을 보완하면 재평가 가능할지 구체 액션 제시
- 문장 길이: 5~8문장
출력은 순수 텍스트만.
"""


def read_docx_text(path: Path) -> str:
    try:
        doc = Document(str(path))
        lines = [p.text.strip() for p in doc.paragraphs if p.text and p.text.strip()]
        return "\n".join(lines)
    except Exception:
        return ""


def call_openai(user_prompt: str) -> tuple[str, dict]:
    api_key = os.getenv("OPENAI_API_KEY", "")
    model = os.getenv("OPENAI_MODEL", "gpt-4o-mini")
    base_url = os.getenv("OPENAI_BASE_URL", "https://api.openai.com/v1")
    if not api_key:
        raise RuntimeError("OPENAI_API_KEY가 설정되지 않았습니다.")

    url = f"{base_url.rstrip('/')}/chat/completions"
    headers = {"Authorization": f"Bearer {api_key}", "Content-Type": "application/json"}
    body = {
        "model": model,
        "temperature": 0.2,
        "response_format": {"type": "json_object"},
        "messages": [
            {"role": "system", "content": SYSTEM_PROMPT},
            {"role": "user", "content": user_prompt},
        ],
    }
    r = requests.post(url, headers=headers, json=body, timeout=120)
    r.raise_for_status()
    data = r.json()
    content = data["choices"][0]["message"]["content"]
    usage = data.get("usage", {})
    return content, usage


def call_gauss(user_prompt: str) -> tuple[str, dict]:
    # 환경에 맞게 .env에서 설정
    api_key = os.getenv("GAUSS_API_KEY", "")
    model = os.getenv("GAUSS_MODEL", "")
    endpoint = os.getenv("GAUSS_ENDPOINT", "")
    if not api_key or not endpoint:
        raise RuntimeError("GAUSS_ENDPOINT/GAUSS_API_KEY가 설정되지 않았습니다.")

    headers = {"Authorization": f"Bearer {api_key}", "Content-Type": "application/json"}
    body = {
        "model": model,
        "messages": [
            {"role": "system", "content": SYSTEM_PROMPT},
            {"role": "user", "content": user_prompt},
        ],
        "temperature": 0.2,
    }
    r = requests.post(endpoint, headers=headers, json=body, timeout=120)
    r.raise_for_status()
    data = r.json()
    # OpenAI 유사 응답/커스텀 응답 둘 다 대응
    usage = data.get("usage", {}) if isinstance(data, dict) else {}
    if "choices" in data:
        return data["choices"][0]["message"]["content"], usage
    if "output_text" in data:
        return data["output_text"], usage
    return json.dumps(data, ensure_ascii=False), usage


def parse_result(raw: str) -> dict:
    try:
        return json.loads(raw)
    except Exception:
        m = re.search(r"\{[\s\S]*\}", raw)
        if not m:
            raise
        return json.loads(m.group(0))


def call_llm_text(provider: str, user_prompt: str) -> tuple[str, dict]:
    if provider == "openai":
        api_key = os.getenv("OPENAI_API_KEY", "")
        model = os.getenv("OPENAI_MODEL", "gpt-4o-mini")
        base_url = os.getenv("OPENAI_BASE_URL", "https://api.openai.com/v1")
        if not api_key:
            raise RuntimeError("OPENAI_API_KEY가 설정되지 않았습니다.")
        url = f"{base_url.rstrip('/')}/chat/completions"
        headers = {"Authorization": f"Bearer {api_key}", "Content-Type": "application/json"}
        body = {
            "model": model,
            "temperature": 0.2,
            "messages": [
                {"role": "system", "content": REJECTION_PROMPT},
                {"role": "user", "content": user_prompt},
            ],
        }
        r = requests.post(url, headers=headers, json=body, timeout=120)
        r.raise_for_status()
        data = r.json()
        return data["choices"][0]["message"]["content"], data.get("usage", {})

    # gauss
    api_key = os.getenv("GAUSS_API_KEY", "")
    model = os.getenv("GAUSS_MODEL", "")
    endpoint = os.getenv("GAUSS_ENDPOINT", "")
    if not api_key or not endpoint:
        raise RuntimeError("GAUSS_ENDPOINT/GAUSS_API_KEY가 설정되지 않았습니다.")
    headers = {"Authorization": f"Bearer {api_key}", "Content-Type": "application/json"}
    body = {
        "model": model,
        "messages": [
            {"role": "system", "content": REJECTION_PROMPT},
            {"role": "user", "content": user_prompt},
        ],
        "temperature": 0.2,
    }
    r = requests.post(endpoint, headers=headers, json=body, timeout=120)
    r.raise_for_status()
    data = r.json()
    usage = data.get("usage", {}) if isinstance(data, dict) else {}
    if "choices" in data:
        return data["choices"][0]["message"]["content"], usage
    if "output_text" in data:
        return data["output_text"], usage
    return json.dumps(data, ensure_ascii=False), usage


def grade_or_default(v: str) -> str:
    valid = {"A+", "A", "B+", "B", "C"}
    return v if v in valid else "B"


def normalize_level(item: str, level: str) -> str:
    v = str(level).strip()
    allowed = set(RUBRIC[item].keys())
    if v in allowed:
        return v
    # 기술 선정 이유는 보통 불가
    if item == "기술 선정 이유":
        return "양호"
    return "보통"


def calc_total_from_rubric_levels(rubric_levels: dict) -> float:
    total = 0.0
    for item, level_map in RUBRIC.items():
        level = normalize_level(item, rubric_levels.get(item, "보통"))
        total += level_map.get(level, 0)
    return round(total, 1)


def score_to_grade(score: float) -> str:
    # 5점 환산 기준
    if score >= 4.5:
        return "A+"
    if score >= 3.8:
        return "A"
    if score >= 3.0:
        return "B+"
    if score >= 2.0:
        return "B"
    return "C"


def build_grades_from_rubric(rubric_levels: dict) -> dict:
    # 19개 세부항목 -> 엑셀 10개 항목 등급으로 집계
    grades = {}
    max_score_per_item = {
        k: max(v.values()) for k, v in RUBRIC.items()
    }
    for metric, items in METRIC_TO_RUBRIC.items():
        earned = 0.0
        maximum = 0.0
        for item in items:
            level = normalize_level(item, rubric_levels.get(item, "보통"))
            earned += RUBRIC[item].get(level, 0)
            maximum += max_score_per_item[item]
        norm5 = (earned / maximum) * 5.0 if maximum else 0.0
        grades[metric] = score_to_grade(norm5)
    return grades


def build_row_index(ws):
    """엑셀의 기존 행(접수번호 B열)과 파일번호를 매핑"""
    row_by_no = {}
    for r in range(5, ws.max_row + 1):
        v = ws.cell(r, 2).value
        if v is None:
            continue
        try:
            no = int(v)
            row_by_no[no] = r
        except Exception:
            continue
    return row_by_no


def write_filtered_sheet(wb, src_ws, title: str, target_decision: str):
    if title in wb.sheetnames:
        del wb[title]
    ws = wb.create_sheet(title)

    # 헤더(4행) 복사
    max_c = src_ws.max_column
    for c in range(1, max_c + 1):
        ws.cell(1, c).value = src_ws.cell(4, c).value

    out_r = 2
    for r in range(5, src_ws.max_row + 1):
        if src_ws.cell(r, 9).value == target_decision:
            for c in range(1, max_c + 1):
                ws.cell(out_r, c).value = src_ws.cell(r, c).value
            out_r += 1


def run_evaluation(zip_file, provider: str, max_docs: int, ui: dict) -> tuple[Path, int, dict]:
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    run_dir = WORK_DIR / f"run_{ts}"
    run_dir.mkdir(parents=True, exist_ok=True)

    zip_path = run_dir / zip_file.name
    zip_path.write_bytes(zip_file.getbuffer())
    extract_dir = run_dir / zip_path.stem
    extract_dir.mkdir(exist_ok=True)

    log_lines = []

    def push_log(msg: str):
        log_lines.append(msg)
        ui["log"].markdown("\n".join([f"- {x}" for x in log_lines[-18:]]))

    push_log(f"📦 ZIP 저장: {zip_path}")
    with zipfile.ZipFile(zip_path, "r") as zf:
        zf.extractall(extract_dir)
    push_log(f"✅ 압축 해제 완료: {extract_dir}")

    docx_files = sorted(extract_dir.rglob("*.docx"))[:max_docs]
    if not docx_files:
        raise RuntimeError("ZIP 내부에 docx 파일이 없습니다.")

    # 템플릿: PoC/평가표.xlsx 우선, 없으면 src 원본 사용
    template = BASE_DIR / "평가표.xlsx"
    if not template.exists():
        raise RuntimeError("PoC/평가표.xlsx 템플릿이 없습니다.")

    wb = openpyxl.load_workbook(template)
    ws = wb[wb.sheetnames[0]]
    row_by_no = build_row_index(ws)

    success = 0
    pass_count = 0
    improve_count = 0
    scored_rows = []  # [{row, doc_no, file, raw_total, final_total, first_round_x}]
    x_penalty_count = 0
    rejection_rewrite_count = 0
    penalty_points = int(os.getenv("FIRST_ROUND_X_PENALTY", "8"))
    token_prompt = 0
    token_completion = 0
    token_total = 0
    started = time.time()

    for i, docx in enumerate(docx_files, start=0):
        txt = read_docx_text(docx)
        txt = txt[:12000]  # 토큰 폭주 방지

        ui["progress"].progress(i / len(docx_files), text=f"평가 준비 중... ({i+1}/{len(docx_files)})")
        push_log(f"🤖 평가 시작 ({i+1}/{len(docx_files)}): {docx.name}")
        ui["status"].info(f"LLM 분석 요청 중: `{docx.name}`")
        prompt = f"파일명: {docx.name}\n\n제안서 본문:\n{txt}"

        if provider == "openai":
            raw, usage = call_openai(prompt)
        else:
            # gauss trial: 분당 3회 제한 고려
            if i > 0:
                time.sleep(int(os.getenv("GAUSS_MIN_INTERVAL_SEC", "22")))
            raw, usage = call_gauss(prompt)

        token_prompt += int(usage.get("prompt_tokens", 0) or 0)
        token_completion += int(usage.get("completion_tokens", 0) or 0)
        token_total += int(usage.get("total_tokens", 0) or 0)

        data = parse_result(raw)
        rubric_levels = data.get("rubric_levels", {})
        opinion = data.get("opinion", "의견 생성 실패")

        if not rubric_levels:
            rubric_levels = {k: ("양호" if k == "기술 선정 이유" else "보통") for k in RUBRIC.keys()}

        # 파일명 접수번호 기준으로 기존 행에 덮어쓰기 (다른 행 생성 금지)
        doc_no = None
        try:
            doc_no = int(docx.name.split("_")[0])
        except Exception:
            pass

        row = row_by_no.get(doc_no) if doc_no is not None else None
        if row is None:
            push_log(f"⚠️ 행 매핑 실패(건너뜀): {docx.name}")
            continue

        # 기본 컬럼 채우기
        ws.cell(row, 1).value = ws.cell(row, 1).value or "1. MX"
        ws.cell(row, 2).value = doc_no
        ws.cell(row, 5).value = docx.stem
        ws.cell(row, 8).value = opinion

        # 19개 세부항목(우수/양호/보통/미흡) 기록
        for item, c in RUBRIC_COLUMNS.items():
            level = normalize_level(item, rubric_levels.get(item, "보통"))
            ws.cell(row, c).value = level

        # 총점(J열) 계산 + 1차(X) 패널티 반영
        raw_total = calc_total_from_rubric_levels(rubric_levels)

        first_round_val = ws.cell(row, 7).value  # 1차 심사 결과(첨부심사)
        # 일부 파일에서 J컬럼에 X가 들어온 경우도 보조 인식
        alt_val = ws.cell(row, 10).value
        first_round_x = str(first_round_val).strip().upper() == "X" or str(alt_val).strip().upper() == "X"

        final_total = max(0, raw_total - penalty_points) if first_round_x else raw_total
        if first_round_x:
            x_penalty_count += 1

        ws.cell(row, 10).value = final_total

        success += 1
        ui["progress"].progress((i + 1) / len(docx_files), text=f"평가 진행률 {(i + 1)}/{len(docx_files)}")
        ui["status"].success(f"완료: `{docx.name}` | 총점: **{final_total}**")

        penalty_note = f" (X 패널티 -{penalty_points})" if first_round_x else ""
        ui["latest"].markdown(
            f"### 최신 평가 결과\n"
            f"- 파일: `{docx.name}`\n"
            f"- 접수번호: `{doc_no}`\n"
            f"- 총점: **{final_total}점**{penalty_note} (판정 대기)\n"
            f"- 의견 요약: {opinion[:160]}{'...' if len(opinion) > 160 else ''}"
        )

        scored_rows.append({
            "row": row,
            "파일": docx.name,
            "접수번호": doc_no,
            "총점": final_total,
            "원점수": raw_total,
            "1차X": first_round_x,
            "의견": opinion,
            "rubric_levels": rubric_levels,
        })

        if "rows" not in st.session_state:
            st.session_state["rows"] = []
        st.session_state["rows"].append({
            "파일": docx.name,
            "접수번호": doc_no,
            "총점": final_total,
            "1차X": "Y" if first_round_x else "N",
            "판정": "대기",
        })
        ui["table"].dataframe(st.session_state["rows"], use_container_width=True, height=220)
        push_log(f"   └ 완료: {docx.name} (총점 {final_total}{', X패널티' if first_round_x else ''})")

    # 최종 판정: 92/262 비율로 상위 점수 합격
    pass_ratio = 92 / 262
    pass_quota = max(1, round(len(scored_rows) * pass_ratio)) if scored_rows else 0
    ranked = sorted(scored_rows, key=lambda x: x["총점"], reverse=True)

    for idx, item in enumerate(ranked):
        decision = "합격" if idx < pass_quota else "불합격"
        ws.cell(item["row"], 9).value = decision  # I열 최종판정
        if decision == "합격":
            pass_count += 1
        else:
            improve_count += 1
            # 불합격 사유를 LLM으로 상세 재작성
            try:
                if provider == "gauss":
                    time.sleep(int(os.getenv("GAUSS_MIN_INTERVAL_SEC", "22")))
                rl = item.get("rubric_levels", {})
                detail_prompt = (
                    f"파일명: {item['파일']}\n"
                    f"최종점수: {item['총점']} (원점수 {item.get('원점수')}, 1차X={item.get('1차X')})\n"
                    f"기존 의견: {item.get('의견','')}\n"
                    f"세부평가: {json.dumps(rl, ensure_ascii=False)}\n"
                )
                detailed, u2 = call_llm_text(provider, detail_prompt)
                ws.cell(item["row"], 8).value = detailed.strip()
                rejection_rewrite_count += 1
                token_prompt += int(u2.get("prompt_tokens", 0) or 0)
                token_completion += int(u2.get("completion_tokens", 0) or 0)
                token_total += int(u2.get("total_tokens", 0) or 0)
            except Exception:
                # 실패 시 기존 의견 유지
                pass

    # 요약 테이블 최종 반영
    st.session_state["rows"] = []
    for idx, item in enumerate(ranked):
        d = "합격" if idx < pass_quota else "불합격"
        st.session_state["rows"].append({
            "파일": item["파일"],
            "접수번호": item["접수번호"],
            "총점": item["총점"],
            "판정": d,
        })
    ui["table"].dataframe(st.session_state["rows"], use_container_width=True, height=220)

    # 합격/불합격 시트 생성
    write_filtered_sheet(wb, ws, "합격", "합격")
    write_filtered_sheet(wb, ws, "불합격", "불합격")

    out_path = BASE_DIR / f"평가표_{ts}.xlsx"
    wb.save(out_path)
    elapsed_sec = round(time.time() - started, 2)
    ui["progress"].progress(1.0, text="평가 완료")
    ui["status"].success(f"전체 완료: {success}/{len(docx_files)}건 저장")
    stats = {
        "evaluated": success,
        "pass_count": pass_count,
        "improve_count": improve_count,
        "x_penalty_count": x_penalty_count,
        "rejection_rewrite_count": rejection_rewrite_count,
        "penalty_points": penalty_points,
        "elapsed_sec": elapsed_sec,
        "prompt_tokens": token_prompt,
        "completion_tokens": token_completion,
        "total_tokens": token_total,
    }
    return out_path, success, stats


def main():
    st.set_page_config(page_title="제안서 자동 평가 에이전트 (PoC)", layout="wide")
    st.title("제안서 자동 평가 에이전트 (PoC)")

    c1, c2, c3 = st.columns([1, 1.2, 1])

    with c1:
        st.subheader("입력")
        provider = st.selectbox("LLM 선택", ["openai", "gauss"])
        max_docs = st.number_input("평가 문서 수", min_value=1, max_value=20, value=5, step=1)
        zip_file = st.file_uploader("제안서.zip 업로드", type=["zip"])
        run_btn = st.button("평가 시작", type="primary", use_container_width=True)

    with c2:
        st.subheader("처리")
        process_area = st.container(border=True)
        process_area.caption("LLM 호출/파일 처리 로그")
        status_box = st.empty()
        progress_bar = st.progress(0, text="대기 중")
        latest_box = st.empty()
        table_box = st.empty()
        log_box = st.empty()

    with c3:
        st.subheader("출력")
        result_area = st.container(border=True)

    if run_btn:
        if not zip_file:
            st.error("ZIP 파일을 먼저 업로드해 주세요.")
            return

        st.session_state["rows"] = []
        status_box.info("🚀 작업 시작...")
        time.sleep(0.3)

        ui = {
            "status": status_box,
            "progress": progress_bar,
            "latest": latest_box,
            "table": table_box,
            "log": log_box,
        }

        try:
            out_path, n, stats = run_evaluation(zip_file, provider, int(max_docs), ui)
            with c3:
                st.success("평가 결과 분석 완료")
                st.markdown(
                    "### 평가 결과 요약 (상세)\n"
                    f"- 평가 건수: **{stats['evaluated']}건**\n"
                    f"- 합격권: **{stats['pass_count']}건**\n"
                    f"- 불합격: **{stats['improve_count']}건**\n"
                    f"- 1차 X 패널티 적용: **{stats['x_penalty_count']}건** (건당 -{stats['penalty_points']}점)\n"
                    f"- 불합격 사유 상세 재작성: **{stats['rejection_rewrite_count']}건**\n"
                    f"- 처리 시간: **{stats['elapsed_sec']}초**\n"
                    f"- 총 사용 토큰: **{stats['total_tokens']}** "
                    f"(입력 {stats['prompt_tokens']} / 출력 {stats['completion_tokens']})"
                )
                st.write(f"결과 파일: `{out_path.name}`")
                st.download_button(
                    "엑셀 다운로드",
                    data=out_path.read_bytes(),
                    file_name=out_path.name,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                )
        except Exception as e:
            with c3:
                st.error(f"오류: {e}")


if __name__ == "__main__":
    main()
